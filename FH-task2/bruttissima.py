#This code returns 4 sheets in which for every base, the amount of fleet hours is counted.
from typing import List, Any
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import numpy as np


# Funzione che calcola i dati e restituisce una lista di tuple
def calcola_dati():
    return [
        (0, "A", 10),
        (1, "A", 12),
        (0, "B", 20),
        (1, "B", 18),
        (0, "C", 15),
        (1, "C", 17),
        (0, "D", 25),
        (1, "D", 22),
    ]


# Funzione per generare il barplot
def genera_barplot(dati):
    # Estrarre le categorie uniche
    categorie = sorted(set(cat for _, cat, _ in dati))

    # Creare dizionari per i valori
    valori_0 = {cat: 0 for cat in categorie}  # Tipo 0
    valori_1 = {cat: 0 for cat in categorie}  # Tipo 1

    # Riempire i dizionari con i dati
    for tipo, categoria, valore in dati:
        if tipo == 0:
            valori_0[categoria] = valore
        else:
            valori_1[categoria] = valore

    # Convertire i dizionari in liste ordinate
    valori_0 = [valori_0[cat] for cat in categorie]
    valori_1 = [valori_1[cat] for cat in categorie]

    # Posizioni per le barre
    x = np.arange(len(categorie))  # Indici delle categorie
    width = 0.3  # Larghezza delle barre

    # Creazione del barplot
    fig, ax = plt.subplots()
    ax.bar(x - width / 2, valori_0, width, label="Tipo 0", color="blue")
    ax.bar(x + width / 2, valori_1, width, label="Tipo 1", color="red")

    # Etichette e titoli
    ax.set_xlabel("Categorie")
    ax.set_ylabel("Valori")
    ax.set_title("Barplot con colonne affiancate per Tipo")
    ax.set_xticks(x)
    ax.set_xticklabels(categorie)
    ax.legend()

    # Mostrare il grafico
    plt.show()


# Chiamata alle funzioni
dati = calcola_dati()
genera_barplot(dati)

#####################################################################


# def generate_excel_fh_bases(file_input, output_file):
#     df_aircraft_base_position = pd.read_excel(file_input, sheet_name='aircraft_base_position', index_col=0)
#     df_fh = pd.read_excel(file_input, sheet_name='FH', index_col=0)
#
#     if 'Total' in df_fh.columns:
#         df_fh = df_fh.drop(columns=['Total'])
#
#     if 'Totale AC' in df_fh.columns:
#         df_fh = df_fh.drop(columns=['Totale AC'])
#
#     if 'Totale AT' in df_fh.columns:
#         df_fh = df_fh.drop(columns=['Totale AT'])
#
#     if 'Gen.1' in df_fh.columns:
#         df_fh = df_fh.drop(columns=['Gen.1'])
#
#     # List of all the bases and months in the excel
#     months = df_fh.columns[1:]  # Tutti i mesi tranne il primo
#     bases = df_aircraft_base_position.values[:, 1:]  # Tutte le basi, saltando la prima colonna
#
#     # Group the unique bases of the sheet aircraft_base_position
#     basi_uniche = sorted(set(bases.flatten()))
#
#     # Dictionary that contains all the Dataframe for each base
#     dizionario_output = {}
#
#     for base in basi_uniche:
#         # Create for each base a list of the aircrafts that flew in at least once during the year
#         aircrafts_in_base = df_aircraft_base_position[df_aircraft_base_position.eq(base).any(axis=1)].index.tolist()
#
#         # Create an empty dataframe for each base
#         output_df = pd.DataFrame(index=aircrafts_in_base, columns=df_fh.columns)
#         output_df.index.name = 'aircraft'
#
#         # Populate the dataframe with the data of the FH sheet for each base
#         for aircraft in aircrafts_in_base:
#             if aircraft in df_fh.index:
#                 # Extract the corresponding row of the sheet
#                 fh_row = df_fh.loc[aircraft]
#                 # This is a cycle that iterates for each month in order to check if the aircraft flew in the base of interest
#                 for month in df_fh.columns:
#                     if month in df_aircraft_base_position.columns:  # check that the month appears in both sheet
#                         if df_aircraft_base_position.loc[aircraft, month] == base:
#                             # If the aircraft flew in the corrispondig base, add to the dataframe the corrisponding row of the fh sheet
#                             output_df.loc[aircraft, month] = fh_row[month]
#                         else:
#                             # If aircraft didn't fly on that base we add a "-"
#                             output_df.loc[aircraft, month] = "-"
#             else:
#                 print(f"Attenzione: {aircraft} non trovato in df_fh")
#
#         # Add the DataFrame to the dictionary
#         dizionario_output[base] = output_df
#
#     # Create an Excel output file with all the sheets
#     with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
#         for base, df in dizionario_output.items():
#             nome_foglio = f"{base}_FH"  # Rename the sheet
#             df.to_excel(writer, sheet_name=nome_foglio, index=True, header=True)
#
#     # Remove the bold
#     wb = load_workbook(output_file)
#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         for cell in ws[1]:  # Prima riga (header)
#             cell.font = Font(bold=False)
#         for row in ws.iter_rows(min_row=2):  # Colonna A (index)
#             row[0].font = Font(bold=False)
#
#     wb.save(output_file)
#     print(f"File Excel in {output_file}")
#
# ###################################################################
#
# file_input = "C:/Users/Utente/Desktop/Tesi/simulations/NEWDELAY/pafam_optimization_results_anno5_bayes_NEWDELAY.xlsx"
# output_file = 'output file name.xlsx'
# generate_excel_fh_bases(file_input, output_file)